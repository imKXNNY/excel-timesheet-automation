# src/import_employee_data.py
import os
import openpyxl
from datetime import datetime

from src.logger import setup_logging
setup_logging()
import logging

def import_employee_data(admin_file, employee_files_dir):
    try:
        # Öffne das Admin-Dashboard
        admin_wb = openpyxl.load_workbook(admin_file)
        if 'Dashboard' not in admin_wb.sheetnames:
            admin_ws = admin_wb.create_sheet(title='Dashboard')
            headers = ['Mitarbeitername', 'Gesamtstunden', 'Überstunden', 'Urlaubstage', 'Status']
            for col, header in enumerate(headers, start=1):
                admin_ws.cell(row=1, column=col, value=header)
            logging.info("Dashboard-Sheet erstellt und Header hinzugefügt.")
        else:
            admin_ws = admin_wb['Dashboard']

        # Iteriere durch alle Mitarbeiter-Dateien im Verzeichnis
        for filename in os.listdir(employee_files_dir):
            if filename.endswith('.xlsx') and filename != os.path.basename(admin_file):
                employee_file = os.path.join(employee_files_dir, filename)
                try:
                    employee_wb = openpyxl.load_workbook(employee_file)
                except Exception as e:
                    logging.error(f"Fehler beim Öffnen der Datei {filename}: {e}")
                    continue

                if 'Angepasste Arbeitszeiterfassung' not in employee_wb.sheetnames:
                    logging.warning(f"'Angepasste Arbeitszeiterfassung' Sheet nicht gefunden in {filename}")
                    continue

                employee_ws = employee_wb['Angepasste Arbeitszeiterfassung']

                # Extrahiere den Mitarbeiternamen
                employee_name = employee_ws['B3'].value
                month_year = employee_ws['B4'].value

                logging.info(f"Importiere Daten für Mitarbeiter: {employee_name}, Monat/Jahr: {month_year}")

                # Berechne Gesamtstunden und Überstunden
                total_hours = 0
                overtime = 0
                for row in range(7, employee_ws.max_row + 1):
                    hours = employee_ws.cell(row=row, column=5).value
                    if hours:
                        total_hours += hours
                        if hours > 8:
                            overtime += hours - 8

                # Berechne Urlaubstage (Tage ohne Arbeitsstunden)
                vacation_days = 0
                for row in range(7, employee_ws.max_row + 1):
                    date_cell = employee_ws.cell(row=row, column=1).value
                    if isinstance(date_cell, datetime):
                        date = date_cell
                    else:
                        try:
                            date = datetime.strptime(str(date_cell), "%Y-%m-%d %H:%M:%S")
                        except ValueError:
                            continue

                    # Annahme: Tage ohne Arbeitsstunden sind Urlaubstage
                    hours = employee_ws.cell(row=row, column=5).value
                    if hours == 0 or hours is None:
                        if date.weekday() < 5:  # Montag bis Freitag
                            vacation_days += 1

                # Aktualisiere das Admin-Dashboard
                updated = False
                for row in range(2, admin_ws.max_row + 1):
                    if admin_ws.cell(row=row, column=1).value == employee_name:
                        admin_ws.cell(row=row, column=2, value=total_hours)
                        admin_ws.cell(row=row, column=3, value=overtime)
                        admin_ws.cell(row=row, column=4, value=vacation_days)
                        admin_ws.cell(row=row, column=5, value="Aktualisiert")
                        updated = True
                        logging.info(f"Daten für {employee_name} aktualisiert.")
                        break

                if not updated:
                    new_row = admin_ws.max_row + 1
                    admin_ws.cell(row=new_row, column=1, value=employee_name)
                    admin_ws.cell(row=new_row, column=2, value=total_hours)
                    admin_ws.cell(row=new_row, column=3, value=overtime)
                    admin_ws.cell(row=new_row, column=4, value=vacation_days)
                    admin_ws.cell(row=new_row, column=5, value="Neu hinzugefügt")
                    logging.info(f"Mitarbeiter {employee_name} neu hinzugefügt.")

        # Speichere das aktualisierte Admin-Dashboard
        admin_wb.save(admin_file)
        logging.info(f"Admin-Dashboard wurde aktualisiert: {admin_file}")

    except Exception as e:
        logging.error(f"Fehler beim Importieren der Mitarbeiterdaten: {e}")
