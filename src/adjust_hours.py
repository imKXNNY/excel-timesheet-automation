# src/adjust_hours.py
import os
import openpyxl
from datetime import datetime
import math

from src.logger import setup_logging
setup_logging()
import logging

def extract_month_year_from_dates(ws):
    """
    Extrahiert das Monat/Jahr aus den Datumseinträgen ab Zeile 7.
    Annahme: Alle Einträge befinden sich im gleichen Monat.
    """
    for row in range(7, ws.max_row + 1):
        date = ws.cell(row=row, column=1).value
        if isinstance(date, datetime):
            return date.strftime("%B %Y")
        else:
            try:
                date = datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S")
                return date.strftime("%B %Y")
            except ValueError:
                continue
    logging.error("Keine gültigen Datumseinträge gefunden.")
    raise ValueError("Keine gültigen Datumseinträge gefunden.")

def calculate_total_hours(start_time, end_time, pause_minutes):
    """
    Berechnet die Gesamtarbeitsstunden für einen Tag.
    
    Args:
        start_time (datetime): Startzeit des Arbeitstages.
        end_time (datetime): Endzeit des Arbeitstages.
        pause_minutes (float): Pausenzeit in Minuten.
        
    Returns:
        float: Gesamtarbeitsstunden, gerundet auf die nächsten 15 Minuten.
    """
    if end_time <= start_time:
        logging.warning("Endzeit liegt nicht nach Startzeit.")
        return 0
    
    # Berechnung der Gesamtstunden abzüglich der Pause
    total_hours = (end_time - start_time).total_seconds() / 3600 - (pause_minutes / 60)
    
    if total_hours < 0:
        logging.warning("Pausenzeit übersteigt die Arbeitszeit.")
        total_hours = 0
    
    # Rundung auf die nächsten 15 Minuten
    total_hours = math.ceil(total_hours * 4) / 4
    return total_hours

def extract_info_from_filename(filename):
    """
    Extrahiert den Mitarbeiternamen und das Monat/Jahr aus dem Dateinamen.
    
    Erwartetes Format: Mitarbeitername_Jahr-Monat.xlsx
    Beispiel: Gregor_Modavski_2024-04.xlsx
    """
    base = os.path.basename(filename)
    name_part = os.path.splitext(base)[0]  # Entfernt die Dateiendung
    parts = name_part.split('_')
    
    if len(parts) < 2:
        logging.error("Dateiname entspricht nicht dem erwarteten Format.")
        raise ValueError("Dateiname entspricht nicht dem erwarteten Format.")
    
    employee_name = '_'.join(parts[:-1])  # Alle Teile außer dem letzten sind der Name
    month_year_str = parts[-1]
    
    try:
        month_year = datetime.strptime(month_year_str, "%Y-%m")
        month_year_formatted = month_year.strftime("%B %Y")  # z.B., April 2024
    except ValueError:
        logging.error("Monat/Jahr im Dateinamen ist nicht korrekt formatiert.")
        raise ValueError("Monat/Jahr im Dateinamen ist nicht korrekt formatiert.")
    
    return employee_name, month_year_formatted

def adjust_hours(input_file, output_file):
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        # Extrahiere Informationen aus dem Dateinamen
        employee_name, month_year = extract_info_from_filename(input_file)
        logging.info(f"Verarbeite Daten für Mitarbeiter: {employee_name}, Monat/Jahr: {month_year}")

        total_hours = 0
        weekly_hours = []
        current_week_hours = 0
        week_start = None
        overtime_pool = 0

        # Durchlaufe alle Zeilen und berechne Arbeitsstunden
        for row in range(7, ws.max_row + 1):
            date = ws.cell(row=row, column=1).value
            if not isinstance(date, datetime):
                try:
                    date = datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    logging.warning(f"Ungültiges Datum in Zeile {row}: {date}")
                    continue

            # Lese Arbeitsbeginn, Arbeitsende und Pausenzeit
            start_time = ws.cell(row=row, column=2).value
            end_time = ws.cell(row=row, column=3).value
            pause_minutes = ws.cell(row=row, column=4).value or 0

            # Berechne die Arbeitszeit
            if isinstance(start_time, datetime) and isinstance(end_time, datetime):
                hours = calculate_total_hours(start_time, end_time, pause_minutes)
            else:
                hours = 0  # Wenn ungültige Werte, setze Arbeitszeit auf 0

            # Wöchentliche Stundenberechnung und Überstunden
            if week_start is None or (date - week_start).days >= 7:
                if week_start is not None:
                    weekly_hours.append(current_week_hours)
                week_start = date
                current_week_hours = 0

            current_week_hours += hours
            total_hours += hours

        if week_start is not None:
            weekly_hours.append(current_week_hours)

        # Anpassung der Stunden (z. B. Überstundenmanagement)
        for i, week_hours in enumerate(weekly_hours):
            if week_hours < 39.5:
                deficit = 39.5 - week_hours
                if overtime_pool >= deficit:
                    overtime_pool -= deficit
                    weekly_hours[i] = 39.5
                else:
                    weekly_hours[i] += overtime_pool
                    overtime_pool = 0
            elif week_hours > 39.5:
                overtime = week_hours - 39.5
                overtime_pool += overtime
                weekly_hours[i] = 39.5

        # Verteilung der verbleibenden Überstunden
        for i in range(len(weekly_hours)):
            if overtime_pool > 0:
                added_overtime = min(overtime_pool, 10)  # Max. 10 Überstunden pro Woche
                weekly_hours[i] += added_overtime
                overtime_pool -= added_overtime

        # Erstellen der Ausgabedatei
        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_ws.title = "Angepasste Arbeitszeiterfassung"

        # Kopiere Header und Mitarbeiterinformationen
        for row in range(1, 7):
            for col in range(1, 6):
                output_ws.cell(row=row, column=col, value=ws.cell(row=row, column=col).value)

        # Fülle die angepassten Stunden ein
        current_week = 0
        remaining_hours = weekly_hours[current_week] if weekly_hours else 0
        for row in range(7, ws.max_row + 1):
            date = ws.cell(row=row, column=1).value
            if not isinstance(date, datetime):
                try:
                    date = datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue

            if date.weekday() == 0 and row != 7:  # Neuer Montag, außer der erste Eintrag
                current_week += 1
                if current_week < len(weekly_hours):
                    remaining_hours = weekly_hours[current_week]
                else:
                    remaining_hours = 0

            if remaining_hours > 0:
                hours = min(8, remaining_hours)  # Max. 8 Stunden pro Tag
                remaining_hours -= hours
            else:
                hours = 0

            for col in range(1, 6):
                if col == 5:  # Gesamtstunden
                    output_ws.cell(row=row, column=col, value=hours)
                else:
                    output_ws.cell(row=row, column=col, value=ws.cell(row=row, column=col).value)

        output_wb.save(output_file)
        logging.info(f"Angepasste Arbeitszeiterfassung wurde erstellt: {output_file}")
    except Exception as e:
        logging.error(f"Fehler bei der Analyse der Daten: {e}")
        return "Fehler bei der Datenanalyse."

def analyze_adjusted_data(output_file):
    try:
        wb = openpyxl.load_workbook(output_file)
        ws = wb.active

        employee_name = ws['B3'].value
        month_year = ws['B4'].value

        if not isinstance(employee_name, str) or not isinstance(month_year, str):
            logging.error("Mitarbeitername oder Monat/Jahr sind nicht korrekt ausgefüllt.")
            raise ValueError("Mitarbeitername oder Monat/Jahr sind nicht korrekt ausgefüllt.")

        logging.info(f"Employee Name: {employee_name}")
        logging.info(f"Month/Year: {month_year}")

        analysis = f"Analyse für {employee_name} - {month_year}\n"
        analysis += "-"*50 + "\n"

        total_hours = 0
        overtime_hours = 0
        vacation_days = 0

        for row in range(7, ws.max_row + 1):
            date = ws.cell(row=row, column=1).value
            if not isinstance(date, datetime):
                try:
                    date = datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue

            hours_cell = ws.cell(row=row, column=5).value
            try:
                hours = float(hours_cell) if hours_cell else 0
            except ValueError:
                logging.warning(f"Ungültiger Stundenwert in Zeile {row}: {hours_cell}")
                hours = 0

            total_hours += hours

            if hours > 8:
                overtime_hours += hours - 8

            # Annahme: Tage ohne Arbeitsstunden sind Urlaubstage
            if hours == 0 and date.weekday() < 5:
                vacation_days += 1

        analysis += f"Gesamtstunden: {total_hours}\n"
        analysis += f"Überstunden: {overtime_hours}\n"
        analysis += f"Urlaubstage: {vacation_days}\n"

        logging.info("Datenanalyse abgeschlossen.")
        return analysis

    except Exception as e:
        logging.error(f"Fehler bei der Analyse der Daten: {e}")
        return "Fehler bei der Datenanalyse."