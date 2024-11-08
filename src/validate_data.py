# src/validate_data.py
import openpyxl
from datetime import datetime

from src.logger import setup_logging
setup_logging()
import logging

from src import extract_info_from_filename

def validate_data(input_file, output_file):
    try:
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        # Extrahiere Informationen aus dem Dateinamen
        employee_name, month_year = extract_info_from_filename(input_file)

        logging.info(f"Validiere Daten für Mitarbeiter: {employee_name}, Monat/Jahr: {month_year}")

        # Extrahiere Monat und Jahr als Zahlen
        month, year = map(int, month_year.split())

        last_row = ws.max_row
        errors = []

        for row in range(7, last_row + 1):
            date = ws.cell(row=row, column=1).value
            if isinstance(date, datetime):
                if date.month != month or date.year != year:
                    error_msg = f"Fehler in Zeile {row}: Datum {date} liegt außerhalb des angegebenen Monats/Jahres."
                    errors.append(error_msg)
                    logging.error(error_msg)
            else:
                try:
                    date = datetime.strptime(str(date), "%Y-%m-%d %H:%M:%S")
                    if date.month != month or date.year != year:
                        error_msg = f"Fehler in Zeile {row}: Datum {date} liegt außerhalb des angegebenen Monats/Jahres."
                        errors.append(error_msg)
                        logging.error(error_msg)
                except ValueError:
                    error_msg = f"Fehler in Zeile {row}: Ungültiges Datumformat."
                    errors.append(error_msg)
                    logging.error(error_msg)

            # ... Weitere Validierungen wie zuvor ...

        if errors:
            for error in errors:
                logging.info(error)
            raise ValueError("Datenvalidierung fehlgeschlagen. Überprüfe die Log-Datei für Details.")
        else:
            logging.info("Datenvalidierung abgeschlossen ohne Fehler.")
            print("Datenvalidierung abgeschlossen ohne Fehler.")

    except Exception as e:
        logging.error(f"Fehler bei der Datenvalidierung: {e}")
        print(f"Fehler bei der Datenvalidierung: {e}")
